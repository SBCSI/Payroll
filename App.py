import pandas as pd
import numpy as np
import datetime
#SB
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

# Read the Excel files
df1 = pd.read_excel('C:/test/Test1.xlsx')
df2 = pd.read_excel('C:/test/Test2.xlsx')

# Keep only 'Employee name', 'Ticket Date' and 'Agency' columns in df2
df2 = df2[['Employee Name', 'Ticket Date', 'Agency', 'Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description']]

# Convert 'Ticket Date' to datetime in both dataframes
df1['Ticket Date'] = pd.to_datetime(df1['Ticket Date'])
df2['Ticket Date'] = pd.to_datetime(df2['Ticket Date'])

# If 'Actual Clock In Time' and 'Actual Clock Out Time' are not datetime, convert them
df1['Actual Clock In Time'] = pd.to_datetime(df1['Actual Clock In Time'])
df1['Actual Clock Out Time'] = pd.to_datetime(df1['Actual Clock Out Time'])

# Calculate 'Actual Hours Worked' as the difference between 'Actual Clock Out Time' and 'Actual Clock In Time', converted to hours
df1['Actual Hours Worked'] = (df1['Actual Clock Out Time'] - df1['Actual Clock In Time']).dt.total_seconds() / 3600

# Round 'Actual Hours Worked' to 2 decimal places
df1['Actual Hours Worked'] = df1['Actual Hours Worked'].round(2)

# Merge dataframes based on 'Employee name' and 'Ticket Date'
merged_df = pd.merge(df1, df2, on=['Employee Name', 'Ticket Date'], how='left')

# If 'Agency' is blank, fill with 'CSI'
merged_df['Agency'] = merged_df['Agency'].fillna('CSI')

# Create a new dataframe for rows with errors
errors_df = merged_df[(merged_df['Actual Clock In Time'].isna()) |
                      (merged_df['Actual Clock Out Time'].isna()) |
                      (merged_df['Actual Hours Worked'] < 8)].copy()


# Create the 'Error Description' column
def generate_error_desc(row):
    if pd.isnull(row['Actual Clock In Time']) and pd.isnull(row['Actual Clock Out Time']):
        return 'No Clock In or Clock Out Time'
    elif pd.isnull(row['Actual Clock In Time']):
        return 'No Clock In'
    elif pd.isnull(row['Actual Clock Out Time']):
        return 'No Clock Out'
    elif row['Actual Hours Worked'] < 8:
        return 'Less Than 8 Hours'
    else:
        return np.nan

errors_df['Error Description'] = errors_df.apply(generate_error_desc, axis=1)

# Convert 'Ticket Date' back to 'mm/dd/yyyy' format
merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

# Write the dataframes into a new Excel file with two sheets
with pd.ExcelWriter('C:/test/Payroll.xlsx') as writer:
    merged_df.to_excel(writer, sheet_name='Payroll', index=False)
    errors_df.to_excel(writer, sheet_name='Errors', index=False)

#SB
# Load the workbook
wb = load_workbook('C:/test/Payroll.xlsx')

#SB
# Select the sheets
sheet1 = wb['Payroll']

#SB
# Create a red bold font
red_bold_font = Font(color="FF0000", bold=True)

# Check each cell in column E (5th column) for both sheets
for sheet in [sheet1]:
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            if cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                cell.value = 'Clock In Time?'
                cell.font = red_bold_font
            elif cell.column_letter == 'E' and (cell.value is None or cell.value == ''):
                cell.value = 'Clock Out Time?'
                cell.font = red_bold_font

#SB
#Set column widths
sheet1.column_dimensions['A'].width = 11.26
sheet1.column_dimensions['B'].width = 26.14
sheet1.column_dimensions['C'].width = 31.86
sheet1.column_dimensions['D'].width = 19
sheet1.column_dimensions['E'].width = 20.43
sheet1.column_dimensions['F'].width = 18.71
sheet1.column_dimensions['G'].width = 20.86
sheet1.column_dimensions['H'].width = 32.57
sheet1.column_dimensions['I'].width = 28.71
sheet1.column_dimensions['B'].width = 22.86
sheet1.column_dimensions['K'].width = 57.86



#SB
# Save workbook
wb.save('C:/test/Payroll.xlsx')






